#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""每日自动更新 task-clarify-heuristic 数据追踪表：
拉取 GitHub / ModelScope 公开指标，追加或更新当日行，并重建趋势折线图。

用法（由定时任务或手动调用）：
  python scripts/update_metrics.py
"""
import urllib.request
import json
import datetime
import openpyxl
from openpyxl.chart import LineChart, Reference

FILE = r"C:\Users\罗\WorkBuddy\2026-08-20-15-16-02\task-clarify-heuristic\METRICS_TRACKER.xlsx"
GH_URL = "https://api.github.com/repos/luoxianqiang158/task-clarify-heuristic"
MS_URL = "https://modelscope.cn/api/v1/skills/luoxianqiang/task-clarify-heuristic"


def get_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch():
    gh_s = gh_f = ms_d = ms_l = None
    try:
        d = get_json(GH_URL)
        gh_s = d.get("stargazers_count")
        gh_f = d.get("forks_count")
    except Exception as e:
        print("WARN GitHub:", e)
    try:
        d = get_json(MS_URL).get("Data", {})
        ms_d = d.get("DownloadCount")
        ms_l = d.get("Likes")
    except Exception as e:
        print("WARN ModelScope:", e)
    return gh_s, gh_f, ms_d, ms_l


def rebuild_chart(ws, last_row):
    """清除旧图并基于当前数据重建趋势折线图（列 B~E）。"""
    ws._charts = []
    if last_row < 3:
        return
    ch = LineChart()
    ch.title = "Star / 下载增长趋势"
    ch.style = 12
    ch.y_axis.title = "数量"
    ch.x_axis.title = "日期"
    ch.height = 9
    ch.width = 20
    data = Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=last_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, f"A{last_row + 3}")


def main():
    gh_s, gh_f, ms_d, ms_l = fetch()
    wb = openpyxl.load_workbook(FILE)
    ws = wb["数据追踪"]
    today = datetime.date.today().isoformat()

    last = 2
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            last = r

    if ws.cell(last, 1).value == today:
        target = last          # 当天已存在 → 更新
    else:
        target = last + 1      # 新的一天 → 追加

    ws.cell(target, 1, today)
    if gh_s is not None:
        ws.cell(target, 2, gh_s)
    if gh_f is not None:
        ws.cell(target, 3, gh_f)
    if ms_d is not None:
        ws.cell(target, 4, ms_d)
    if ms_l is not None:
        ws.cell(target, 5, ms_l)
    if target == last + 1:
        ws.cell(target, 10, "自动记录（update_metrics.py）")

    new_last = max(last, target)
    rebuild_chart(ws, new_last)
    wb.save(FILE)
    print(f"row {target} | GitHub star={gh_s} fork={gh_f} | ModelScope dl={ms_d} like={ms_l}")


if __name__ == "__main__":
    main()
