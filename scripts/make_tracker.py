#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成 task-clarify-heuristic 发布后数据追踪表 METRICS_TRACKER.xlsx（含趋势折线图）。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from datetime import date

OUT = "C:/Users/罗/WorkBuddy/2026-08-20-15-16-02/task-clarify-heuristic/METRICS_TRACKER.xlsx"

HEADERS = [
    "日期", "GitHub Star", "GitHub Fork",
    "ModelScope 下载", "ModelScope 点赞",
    "掘金阅读", "掘金点赞",
    "知乎阅读", "知乎赞同",
    "备注",
]

INITIAL = {
    "日期": "2026-08-21",
    "GitHub Star": 1,
    "GitHub Fork": 0,
    "ModelScope 下载": 2,
    "ModelScope 点赞": 1,
    "掘金阅读": None, "掘金点赞": None,
    "知乎阅读": None, "知乎赞同": None,
    "备注": "发布首日基线（GitHub API + ModelScope API 实测）",
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
NOTE_FONT = Font(size=10, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def rebuild_chart(ws, last_row):
    """清除旧图并基于当前数据重建趋势折线图（列 B~E：GitHub Star/Fork、MS 下载/点赞）。"""
    ws._charts = []
    if last_row < 3:
        return
    ch = LineChart()
    ch.title = "Star / 下载增长趋势"
    ch.style = 12
    ch.y_axis.title = "数量"
    ch.x_axis.title = "日期"
    ch.height = 9
    ch.width = 20
    data = Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=last_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, f"A{last_row + 3}")


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "数据追踪"

ws.merge_cells("A1:J1")
ws["A1"] = "task-clarify-heuristic · 发布后数据追踪表"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

r = 3
for c, h in enumerate(HEADERS, start=1):
    v = INITIAL.get(h)
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = BORDER
    cell.alignment = LEFT if h == "备注" else CENTER

widths = [14, 13, 13, 15, 15, 12, 12, 12, 12, 42]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A3"
rebuild_chart(ws, 3)

ws2 = wb.create_sheet("说明")
lines = [
    ("用途", None),
    ("跟踪 task-clarify-heuristic 在两个平台 + 两篇社区帖的发布后表现，沉淀增长数据。", "note"),
    ("", None),
    ("字段来源与获取方式", None),
    ("GitHub Star / Fork：https://api.github.com/repos/luoxianqiang158/task-clarify-heuristic （公开 API，无需登录）", "note"),
    ("ModelScope 下载 / 点赞：https://modelscope.cn/api/v1/skills/luoxianqiang/task-clarify-heuristic", "note"),
    ("掘金阅读 / 点赞：掘金创作者中心后台（需登录，手动填写）", "note"),
    ("知乎阅读 / 赞同：知乎创作中心后台（需登录，手动填写）", "note"),
    ("", None),
    ("更新方式", None),
    ("自动：每日定时任务会调用 scripts/update_metrics.py，拉取 GitHub+ModelScope 并追加/更新当日行、重建折线图。", "note"),
    ("手动：在「数据追踪」表第 4 行起逐行追加新日期的数据；保留第 3 行作为发布首日基线。", "note"),
    ("", None),
    ("目标参考（可自行修订）", None),
    ("第 1 周：GitHub Star ≥ 10，ModelScope 下载 ≥ 20", "note"),
    ("第 1 月：GitHub Star ≥ 50，ModelScope 下载 ≥ 100，掘金阅读 ≥ 1000", "note"),
]
ws2.column_dimensions["A"].width = 115
row = 1
for text, kind in lines:
    cell = ws2.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT if kind == "note" else Font(bold=True, size=11, color="1F3864")
    row += 1

wb.save(OUT)
print("saved:", OUT)
